import csv
import io
import datetime
from django.shortcuts import render, get_object_or_404
from django.contrib import messages
from django.db import transaction
from django.utils.crypto import get_random_string
from openpyxl import load_workbook

from .models import Organization, Client, Vehicle
from .tasks import check_registration_reminders

# Define column mapping aliases for flexibility
CLIENT_FIELDS_MAP = {
    'first_name': ['first_name', 'first', 'firstname', 'client_first_name'],
    'last_name': ['last_name', 'last', 'lastname', 'client_last_name'],
    'middle_name': ['middle_name', 'middle', 'middlename', 'client_middle_name'],
    'email': ['email', 'email_address', 'mail', 'client_email'],
    'phone_number': ['phone_number', 'phone', 'phone_no', 'client_phone_number', 'client_phone'],
    'dob': ['dob', 'date_of_birth', 'birthdate', 'birth_date', 'client_dob'],
    'gender': ['gender', 'sex', 'client_gender'],
    'ssn': ['ssn', 'social_security', 'social_security_number', 'client_ssn'],
    'driver_license': ['driver_license', 'driver_license_number', 'dl', 'dl_number', 'license_number', 'client_dl'],
    'building_no': ['building_no', 'building', 'building_number', 'client_building_no'],
    'street_address': ['street_address', 'street', 'address', 'client_street_address'],
    'apartment': ['apartment', 'apt', 'suite', 'client_apartment'],
    'city': ['city', 'client_city'],
    'state': ['state', 'client_state'],
    'zip_code': ['zip_code', 'zip', 'zipcode', 'client_zip_code', 'client_zip'],
    'county': ['county', 'client_county'],
}

VEHICLE_FIELDS_MAP = {
    'vin': ['vin', 'vin_number', 'vehicle_vin'],
    'vehicle_type': ['vehicle_type', 'type', 'vehicle_type_choice'],
    'plate_number': ['plate_number', 'plate', 'license_plate', 'plate_no'],
    'year': ['year', 'model_year'],
    'make': ['make', 'brand'],
    'model': ['model'],
    'body_type': ['body_type', 'body'],
    'color': ['color', 'colour'],
    'weight': ['weight', 'gross_weight'],
    'fuel_type': ['fuel_type', 'fuel'],
    'cylinders': ['cylinders', 'engine_cylinders'],
    'seats': ['seats', 'seating_capacity'],
    'vehicle_number': ['vehicle_number', 'internal_vehicle_number', 'veh_number'],
    'dl_number': ['dl_number', 'vehicle_dl_number'],
    'registration_effective_date': ['registration_effective_date', 'reg_effective_date'],
    'registration_expiration_date': ['registration_expiration_date', 'reg_expiration_date', 'expiration_date'],
    'plate_type': ['plate_type', 'plate_type_choice'],
    'insurance_company': ['insurance_company', 'insurance', 'carrier'],
    'insurance_policy_number': ['insurance_policy_number', 'policy_number', 'policy_no'],
    'insurance_effective_date': ['insurance_effective_date', 'ins_effective_date'],
    'insurance_expiration_date': ['insurance_expiration_date', 'ins_expiration_date'],
    'is_priority': ['is_priority', 'priority'],
    'odometer_reading': ['odometer_reading', 'odometer'],
    'odometer_status': ['odometer_status'],
    'max_gross_weight': ['max_gross_weight'],
    'num_axles': ['num_axles', 'axles'],
    'owner_name': ['owner_name'],
    'owner_nys_id': ['owner_nys_id', 'owner_id'],
    'owner_dob': ['owner_dob'],
    'co_registrant_name': ['co_registrant_name'],
    'co_registrant_nys_id': ['co_registrant_nys_id'],
    'co_registrant_dob': ['co_registrant_dob'],
    'has_lien': ['has_lien', 'lien'],
    'lienholder_name': ['lienholder_name'],
    'lienholder_address': ['lienholder_address'],
    'lien_filing_code': ['lien_filing_code'],
    'is_leased': ['is_leased', 'lease'],
    'lessor_name': ['lessor_name'],
    'lessor_address': ['lessor_address'],
}

def parse_date(val):
    if not val:
        return None
    if isinstance(val, (datetime.date, datetime.datetime)):
        return val.date() if isinstance(val, datetime.datetime) else val
    val_str = str(val).strip()
    if not val_str:
        return None
    for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%Y/%m/%d', '%m-%d-%Y', '%d-%m-%Y'):
        try:
            return datetime.datetime.strptime(val_str, fmt).date()
        except ValueError:
            continue
    # Fallback iso format check
    try:
        return datetime.date.fromisoformat(val_str)
    except ValueError:
        pass
    return None

def parse_bool(val):
    if val is None:
        return False
    if isinstance(val, bool):
        return val
    val_str = str(val).strip().lower()
    return val_str in ('true', '1', 'yes', 'y', 't')

def extract_fields(row, mapping):
    result = {}
    for field_name, aliases in mapping.items():
        for alias in aliases:
            # Check direct match
            if alias in row:
                result[field_name] = row[alias]
                break
            # Check normalized key match (no spaces/underscores)
            clean_alias = alias.replace("_", "")
            found = False
            for row_key in row.keys():
                if str(row_key).replace("_", "").lower() == clean_alias:
                    result[field_name] = row[row_key]
                    found = True
                    break
            if found:
                break
    return result

def parse_xlsx(file_content):
    wb = load_workbook(filename=file_content, read_only=True, data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    
    # Extract headers and normalize them
    headers = [str(cell).strip().lower().replace(" ", "_") if cell is not None else "" for cell in rows[0]]
    data_rows = []
    for r in rows[1:]:
        if all(cell is None for cell in r):
            continue
        row_dict = {}
        for idx, header in enumerate(headers):
            if header and idx < len(r):
                row_dict[header] = r[idx]
        data_rows.append(row_dict)
    return data_rows

def parse_csv(file_content):
    try:
        decoded = file_content.read().decode('utf-8')
    except UnicodeDecodeError:
        file_content.seek(0)
        decoded = file_content.read().decode('latin-1')
        
    reader = csv.DictReader(io.StringIO(decoded))
    data_rows = []
    for row in reader:
        if not any(row.values()):
            continue
        # Normalize row keys
        normalized_row = {str(k).strip().lower().replace(" ", "_"): v for k, v in row.items() if k is not None}
        data_rows.append(normalized_row)
    return data_rows

def crm_import_view(request):
    """
    Custom view for importing CRM data inside Django Admin.
    Checks authentication and admin privileges automatically.
    """
    organizations = Organization.objects.filter(is_active=True).order_by('name')
    results = None

    if request.method == 'POST':
        org_id = request.POST.get('organization')
        uploaded_file = request.FILES.get('crm_file')
        
        if not org_id or not uploaded_file:
            messages.error(request, "Please select an organization and upload a valid file.")
        else:
            org = get_object_or_404(Organization, id=org_id)
            filename = uploaded_file.name.lower()
            
            try:
                if filename.endswith('.xlsx'):
                    data_rows = parse_xlsx(uploaded_file)
                elif filename.endswith('.csv'):
                    data_rows = parse_csv(uploaded_file)
                else:
                    raise ValueError("Unsupported file format. Please upload a .csv or .xlsx file.")
                
                clients_created = 0
                clients_updated = 0
                vehicles_created = 0
                vehicles_updated = 0
                errors = []
                
                for idx, row in enumerate(data_rows, start=2): # Headers are row 1
                    try:
                        with transaction.atomic():
                            client_data = extract_fields(row, CLIENT_FIELDS_MAP)
                            vehicle_data = extract_fields(row, VEHICLE_FIELDS_MAP)
                            
                            first_name = str(client_data.get('first_name') or '').strip()
                            last_name = str(client_data.get('last_name') or '').strip()
                            
                            if not first_name or not last_name:
                                raise ValueError("Missing required client fields (first_name, last_name)")
                            
                            # Parse dates and normalize options
                            dob = parse_date(client_data.get('dob'))
                            gender = str(client_data.get('gender') or '').strip().lower()
                            if gender not in [c[0] for c in Client.GENDER_CHOICES]:
                                gender = 'prefer_not_to_say'
                            
                            email = str(client_data.get('email') or '').strip() or None
                            phone = str(client_data.get('phone_number') or '').strip()
                            dl = str(client_data.get('driver_license') or '').strip().upper()
                            
                            # Duplicate checking logic for clients
                            client = None
                            client_qs = Client.objects.filter(
                                first_name__iexact=first_name,
                                last_name__iexact=last_name,
                                organization=org
                            )
                            if dl:
                                client = client_qs.filter(driver_license__iexact=dl).first()
                            if not client and email:
                                client = client_qs.filter(email__iexact=email).first()
                            if not client and phone:
                                client = client_qs.filter(phone_number=phone).first()
                            if not client:
                                client = client_qs.first()
                                
                            is_new_client = False
                            if not client:
                                client = Client(
                                    organization=org,
                                    first_name=first_name,
                                    last_name=last_name
                                )
                                is_new_client = True
                            
                            # Update client attributes
                            if 'middle_name' in client_data:
                                client.middle_name = str(client_data['middle_name'] or '').strip()
                            if 'email' in client_data:
                                client.email = email
                            if 'phone_number' in client_data:
                                client.phone_number = phone
                            if dob:
                                client.dob = dob
                            if gender:
                                client.gender = gender
                            if 'ssn' in client_data:
                                client.ssn = str(client_data['ssn'] or '').strip()
                            if dl:
                                client.driver_license = dl
                                
                            # Address fields mapping
                            for addr_f in ['building_no', 'street_address', 'apartment', 'city', 'state', 'zip_code', 'county']:
                                if addr_f in client_data:
                                    val = str(client_data[addr_f] or '').strip()
                                    if addr_f == 'state':
                                        val = val.upper()[:2]
                                        if val not in [s[0] for s in Client.US_STATES]:
                                            val = 'NY'
                                    setattr(client, addr_f, val)
                            
                            client.save()
                            if is_new_client:
                                clients_created += 1
                            else:
                                clients_updated += 1
                                
                            # Handle Vehicle data if VIN is present
                            vin = str(vehicle_data.get('vin') or '').strip().upper()
                            if vin:
                                if len(vin) != 17 or any(c in vin for c in "IOQ"):
                                    raise ValueError(f"Invalid VIN format: '{vin}'. Must be exactly 17 alphanumeric characters and contain no I, O, or Q.")
                                    
                                vehicle = Vehicle.objects.filter(vin=vin).first()
                                is_new_vehicle = False
                                if not vehicle:
                                    auto_vnum = f"VEH-{get_random_string(6, allowed_chars='0123456789')}"
                                    vehicle = Vehicle(
                                        client=client,
                                        vin=vin,
                                        vehicle_number=auto_vnum
                                    )
                                    is_new_vehicle = True
                                else:
                                    vehicle.client = client # Link/Re-link to correct client
                                
                                # Update fields
                                vehicle.year = int(vehicle_data['year']) if vehicle_data.get('year') else None
                                if 'make' in vehicle_data:
                                    vehicle.make = str(vehicle_data['make'] or '').strip()
                                if 'model' in vehicle_data:
                                    vehicle.model = str(vehicle_data['model'] or '').strip()
                                if 'color' in vehicle_data:
                                    vehicle.color = str(vehicle_data['color'] or '').strip()
                                if 'weight' in vehicle_data:
                                    vehicle.weight = str(vehicle_data['weight'] or '').strip()
                                if 'cylinders' in vehicle_data:
                                    vehicle.cylinders = str(vehicle_data['cylinders'] or '').strip()
                                if 'seats' in vehicle_data:
                                    vehicle.seats = str(vehicle_data['seats'] or '').strip()
                                if 'plate_number' in vehicle_data:
                                    vehicle.plate_number = str(vehicle_data['plate_number'] or '').strip()
                                
                                # Choices Normalization
                                vtype = str(vehicle_data.get('vehicle_type') or '').strip().lower()
                                if vtype in [t[0] for t in Vehicle.VEHICLE_TYPES]:
                                    vehicle.vehicle_type = vtype
                                ptype = str(vehicle_data.get('plate_type') or '').strip().lower()
                                if ptype in [t[0] for t in Vehicle.PLATE_TYPES]:
                                    vehicle.plate_type = ptype
                                ftype = str(vehicle_data.get('fuel_type') or '').strip().lower()
                                if ftype in [f[0] for f in Vehicle.FUEL_TYPES]:
                                    vehicle.fuel_type = ftype
                                btype = str(vehicle_data.get('body_type') or '').strip().lower()
                                if btype in [b[0] for b in Vehicle.BODY_TYPES]:
                                    vehicle.body_type = btype
                                    
                                # Parse dates
                                vehicle.registration_effective_date = parse_date(vehicle_data.get('registration_effective_date'))
                                vehicle.registration_expiration_date = parse_date(vehicle_data.get('registration_expiration_date'))
                                vehicle.insurance_effective_date = parse_date(vehicle_data.get('insurance_effective_date'))
                                vehicle.insurance_expiration_date = parse_date(vehicle_data.get('insurance_expiration_date'))
                                
                                # Insurance and booleans
                                if 'insurance_company' in vehicle_data:
                                    vehicle.insurance_company = str(vehicle_data['insurance_company'] or '').strip()
                                if 'insurance_policy_number' in vehicle_data:
                                    vehicle.insurance_policy_number = str(vehicle_data['insurance_policy_number'] or '').strip()
                                vehicle.is_priority = parse_bool(vehicle_data.get('is_priority'))
                                
                                # Tech/MV82 and Lienholder information
                                for extra_f in ['odometer_reading', 'odometer_status', 'max_gross_weight', 'num_axles',
                                                'owner_name', 'owner_nys_id', 'co_registrant_name', 'co_registrant_nys_id',
                                                'lienholder_name', 'lienholder_address', 'lien_filing_code',
                                                'lessor_name', 'lessor_address']:
                                    if extra_f in vehicle_data:
                                        setattr(vehicle, extra_f, str(vehicle_data[extra_f] or '').strip())
                                        
                                vehicle.owner_dob = parse_date(vehicle_data.get('owner_dob'))
                                vehicle.co_registrant_dob = parse_date(vehicle_data.get('co_registrant_dob'))
                                vehicle.has_lien = parse_bool(vehicle_data.get('has_lien'))
                                vehicle.is_leased = parse_bool(vehicle_data.get('is_leased'))
                                
                                vehicle.save()
                                if is_new_vehicle:
                                    vehicles_created += 1
                                else:
                                    vehicles_updated += 1
                                    
                    except Exception as row_err:
                        errors.append({
                            'row': idx,
                            'message': str(row_err)
                        })
                
                # Check for registrations/reminders automation trigger
                if vehicles_created > 0 or vehicles_updated > 0:
                    check_registration_reminders.delay()
                    
                results = {
                    'clients_created': clients_created,
                    'clients_updated': clients_updated,
                    'vehicles_created': vehicles_created,
                    'vehicles_updated': vehicles_updated,
                    'errors': errors
                }
                messages.success(request, "CRM data processing completed successfully.")
            except Exception as e:
                messages.error(request, f"Failed to parse file: {str(e)}")
                
    return render(request, 'admin/crm_import.html', {
        'organizations': organizations,
        'results': results,
        'title': 'CRM Data Import'
    })
