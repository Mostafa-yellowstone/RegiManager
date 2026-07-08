"""Import Email Marketing CRM contacts from CSV or Excel."""

from __future__ import annotations

import csv
import difflib
import io
import re
from dataclasses import dataclass, field
from typing import BinaryIO

from openpyxl import load_workbook

CONTACT_FIELDS = (
    "name",
    "address_line1",
    "address_line2",
    "address_line3",
    "city",
    "state",
    "zip_code",
    "phone",
    "email",
    "website",
    "notes",
)

CONTACT_FIELDS_MAP: dict[str, list[str]] = {
    "name": [
        "name",
        "full_name",
        "fullname",
        "contact_name",
        "client_name",
        "customer_name",
        "customer",
        "contact",
        "company_name",
        "business_name",
        "account_name",
    ],
    "address_line1": [
        "address_line1",
        "address1",
        "address_line_1",
        "street",
        "street_address",
        "street1",
        "mailing_address",
        "address",
        "addr1",
    ],
    "address_line2": [
        "address_line2",
        "address2",
        "address_line_2",
        "suite",
        "apt",
        "apartment",
        "unit",
        "addr2",
    ],
    "address_line3": [
        "address_line3",
        "address3",
        "address_line_3",
        "addr3",
    ],
    "city": ["city", "town", "municipality"],
    "state": ["state", "st", "province", "region"],
    "zip_code": ["zip_code", "zip", "zipcode", "postal_code", "postal", "postcode"],
    "phone": [
        "phone",
        "phone_number",
        "phonenumber",
        "mobile",
        "cell",
        "cellphone",
        "cell_phone",
        "telephone",
        "tel",
        "phone_no",
        "contact_phone",
    ],
    "email": [
        "email",
        "email_address",
        "emailaddress",
        "e_mail",
        "mail",
        "contact_email",
        "email_addr",
    ],
    "website": ["website", "url", "web", "site", "web_site", "homepage"],
    "notes": ["notes", "note", "comments", "comment", "remarks"],
}

FIRST_NAME_ALIASES = frozenset(
    {
        "first_name",
        "firstname",
        "fname",
        "given_name",
        "givenname",
        "first",
    }
)
LAST_NAME_ALIASES = frozenset(
    {
        "last_name",
        "lastname",
        "lname",
        "surname",
        "family_name",
        "familyname",
        "last",
    }
)


@dataclass
class ImportParseResult:
    contacts: list[dict[str, str]] = field(default_factory=list)
    headers: list[str] = field(default_factory=list)
    column_mapping: dict[str, str] = field(default_factory=dict)
    total_rows: int = 0
    skipped_rows: int = 0


def _normalize_key(value) -> str:
    s = str(value or "").strip().lower()
    s = s.replace("\ufeff", "")
    s = re.sub(r"[^a-z0-9]+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    s = s.replace("e_mail", "email")
    return s


def _stringify_cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if value == int(value):
            return str(int(value))
        return str(value).strip()
    if isinstance(value, int):
        return str(value)
    return str(value).strip()


def _alias_lookup() -> dict[str, str]:
    lookup: dict[str, str] = {}
    for field_name, aliases in CONTACT_FIELDS_MAP.items():
        lookup[_normalize_key(field_name)] = field_name
        for alias in aliases:
            lookup[_normalize_key(alias)] = field_name
    return lookup


_ALIAS_TO_FIELD = _alias_lookup()


def _match_header_to_field(header: str) -> str | None:
    norm = _normalize_key(header)
    if not norm:
        return None
    if norm in _ALIAS_TO_FIELD:
        return _ALIAS_TO_FIELD[norm]
    if norm in FIRST_NAME_ALIASES or norm in LAST_NAME_ALIASES:
        return norm

    for alias, field_name in _ALIAS_TO_FIELD.items():
        if len(alias) < 4:
            continue
        if alias in norm or norm in alias:
            return field_name

    alias_keys = list(_ALIAS_TO_FIELD.keys())
    close = difflib.get_close_matches(norm, alias_keys, n=1, cutoff=0.68)
    if close:
        return _ALIAS_TO_FIELD[close[0]]

    field_keys = list(CONTACT_FIELDS_MAP.keys())
    close_field = difflib.get_close_matches(norm, field_keys, n=1, cutoff=0.78)
    if close_field:
        return close_field[0]
    return None


def build_column_mapping(headers: list[str]) -> dict[str, str]:
    """Map original header labels to contact fields (or first/last name tokens)."""
    mapping: dict[str, str] = {}
    for header in headers:
        if not str(header or "").strip():
            continue
        field_name = _match_header_to_field(header)
        if field_name:
            mapping[str(header)] = field_name
    return mapping


def _map_row(raw_row: dict, column_mapping: dict[str, str]) -> dict[str, str]:
    mapped: dict[str, str] = {}
    first_name = ""
    last_name = ""

    for header, value in raw_row.items():
        cell = _stringify_cell(value)
        if not cell:
            continue
        target = column_mapping.get(header)
        if not target:
            continue
        if target in FIRST_NAME_ALIASES:
            first_name = cell
            continue
        if target in LAST_NAME_ALIASES:
            last_name = cell
            continue
        mapped[target] = cell

    if not mapped.get("name"):
        combined = f"{first_name} {last_name}".strip()
        if combined:
            mapped["name"] = combined

    return mapped


def _row_is_importable(mapped: dict[str, str]) -> bool:
    return bool(
        mapped.get("name")
        or mapped.get("email")
        or mapped.get("phone")
    )


def _finalize_contact_row(mapped: dict[str, str]) -> dict[str, str]:
    if not mapped.get("name"):
        mapped["name"] = mapped.get("email") or mapped.get("phone") or "Imported Contact"
    return {field: mapped.get(field, "") for field in CONTACT_FIELDS}


def _rows_from_csv(file_obj: BinaryIO) -> list[dict]:
    raw = file_obj.read()
    if not raw:
        return []
    text = raw.decode("utf-8-sig", errors="replace")
    if not text.strip():
        return []

    delimiter = ","
    try:
        sample = text[:4096]
        delimiter = csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
    except csv.Error:
        delimiter = ","

    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    rows = []
    for row in reader:
        cleaned = {str(k): v for k, v in row.items() if k is not None}
        if any(_stringify_cell(v) for v in cleaned.values()):
            rows.append(cleaned)
    return rows


def _rows_from_xlsx(file_obj: BinaryIO) -> list[dict]:
    wb = load_workbook(file_obj, read_only=True, data_only=True)
    sheet = wb.active
    rows_iter = list(sheet.iter_rows(values_only=True))
    if not rows_iter:
        return []

    header_row_idx = 0
    for idx, row in enumerate(rows_iter[:5]):
        if sum(1 for cell in row if _stringify_cell(cell)) >= 2:
            header_row_idx = idx
            break

    headers = [_stringify_cell(cell) for cell in rows_iter[header_row_idx]]
    data = []
    for row in rows_iter[header_row_idx + 1 :]:
        if not any(_stringify_cell(cell) for cell in row):
            continue
        row_dict = {}
        for i, header in enumerate(headers):
            if not header:
                continue
            if i < len(row):
                row_dict[header] = row[i]
        data.append(row_dict)
    return data


def parse_contact_import_file(uploaded_file) -> ImportParseResult:
    filename = (getattr(uploaded_file, "name", "") or "").lower()
    if hasattr(uploaded_file, "seek"):
        uploaded_file.seek(0)

    if filename.endswith(".csv"):
        raw_rows = _rows_from_csv(uploaded_file)
    elif filename.endswith((".xlsx", ".xlsm")):
        raw_rows = _rows_from_xlsx(uploaded_file)
    else:
        raise ValueError("Upload a .csv or .xlsx file.")

    if not raw_rows:
        return ImportParseResult()

    headers = list(raw_rows[0].keys())
    column_mapping = build_column_mapping(headers)

    contacts: list[dict[str, str]] = []
    skipped = 0
    for row in raw_rows:
        mapped = _map_row(row, column_mapping)
        if not _row_is_importable(mapped):
            skipped += 1
            continue
        contacts.append(_finalize_contact_row(mapped))

    return ImportParseResult(
        contacts=contacts,
        headers=headers,
        column_mapping=column_mapping,
        total_rows=len(raw_rows),
        skipped_rows=skipped,
    )
